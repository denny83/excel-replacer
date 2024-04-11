from flask import Flask, render_template, request, send_file, flash, redirect, url_for
import os
import secrets
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Генерация случайного секретного ключа
secret_key = secrets.token_hex(16)

app = Flask(__name__)
app.secret_key = secret_key

ALLOWED_EXTENSIONS = {'xls', 'xlsx'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload():
    if 'file' not in request.files:
        flash('Файл не выбран')
        return redirect(request.url)
    file = request.files['file']
    if file.filename == '':
        flash('Файл не выбран')
        return redirect(request.url)
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        # Путь к файлу
        file_path = os.path.join(app.root_path, 'uploads', filename)
        file.save(file_path)
        try:
            threshold = int(request.form['threshold'])
            # Загрузка файла Excel
            wb = load_workbook(file_path)
            ws = wb.active

            # Проверка каждой ячейки в активном листе
            for row in ws.iter_rows(min_row=5, min_col=4, max_col=ws.max_column, max_row=ws.max_row):
                for cell in row:
                    # Проверка, что ячейка не в объединенном диапазоне и не в третьем или последнем столбце
                    if cell.column != 3 and cell.column != ws.max_column:
                        if isinstance(cell.value, (int, float)) and cell.value > threshold:
                            # Замена значения ячейки на порог
                            cell.value = threshold
            
            # Проверка каждой строки в активном листе, начиная с пятой строки
            for row in ws.iter_rows(min_row=5, min_col=4, max_col=ws.max_column, max_row=ws.max_row):
                row_sum_formula = f'=SUM({get_column_letter(4)}{row[0].row}:{get_column_letter(ws.max_column - 1)}{row[0].row})'
                ws.cell(row=row[0].row, column=ws.max_column, value=row_sum_formula)

            # Сохранение измененного файла
            wb.save(file_path)

            # Отправка файла пользователю для скачивания
            return send_file(file_path, as_attachment=True, download_name=filename)
        except ValueError:
            flash('Значение порога не является допустимым целым числом')
        except Exception as e:
            flash(f'Произошла ошибка при обработке файла: {str(e)}')
    else:
        flash('Допустимые типы файлов: xls, xlsx')
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True)
